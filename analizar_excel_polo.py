import openpyxl
from datetime import datetime

# Cargar el archivo Excel
wb = openpyxl.load_workbook(r'c:\Users\Usuario\Documents\proyecto\v2\mundoindustrial V-2\mundoindustrial V-1\mundoindustrial\resources\CONTROL DE PISO POLOS (Respuestas).xlsx', data_only=True)

print('=== ANÁLISIS DEL EXCEL - CONTROL DE PISO POLOS ===\n')
print(f'Hojas disponibles: {wb.sheetnames}\n')

# Obtener la hoja REGISTRO
ws = wb['REGISTRO']

# Mostrar encabezados
print('ENCABEZADOS:')
headers = []
for i, cell in enumerate(ws[1], 1):
    headers.append(cell.value)
    print(f'{i}. {cell.value}')

print(f'\nTotal filas (incluyendo encabezado): {ws.max_row}')
print(f'Total columnas: {ws.max_column}\n')

# Mostrar primeras 5 filas de datos
print('PRIMERAS 5 FILAS DE DATOS:')
for row_idx in range(2, min(7, ws.max_row + 1)):
    print(f'\nFila {row_idx}:')
    for col_idx, header in enumerate(headers, 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        print(f'  {header}: {cell_value}')

# Contar filas válidas (con fecha y orden de producción)
print('\n=== ANÁLISIS DE DATOS ===')
filas_validas = 0
filas_invalidas = 0
fecha_col_idx = None
orden_col_idx = None

# Encontrar índices de columnas importantes
for i, h in enumerate(headers, 1):
    if h and 'FECHA' in str(h).upper():
        fecha_col_idx = i
    if h and 'ORDEN' in str(h).upper() and 'PRODUCCIÓN' in str(h).upper():
        orden_col_idx = i

print(f'Columna FECHA: {fecha_col_idx}')
print(f'Columna ORDEN DE PRODUCCIÓN: {orden_col_idx}\n')

for row_idx in range(2, ws.max_row + 1):
    fecha = ws.cell(row=row_idx, column=fecha_col_idx).value if fecha_col_idx else None
    orden = ws.cell(row=row_idx, column=orden_col_idx).value if orden_col_idx else None
    
    if fecha and orden:
        filas_validas += 1
    else:
        filas_invalidas += 1
        if filas_invalidas <= 5:
            print(f'Fila {row_idx} inválida - Fecha: {fecha}, Orden: {orden}')

print(f'\nFILAS VÁLIDAS: {filas_validas}')
print(f'FILAS INVÁLIDAS: {filas_invalidas}')

wb.close()
